#!/usr/bin/env python
# -*- coding: utf-8 -*-
# version 1.1 update by le @ 2017.7.6

import argparse
from argparse import RawTextHelpFormatter
from db import MetaData, DBO
from openpyxl import load_workbook

Filters = {}
Filters["all"] = ''
Filters["default"] = 'and n.等级<>"None" and n.解决办法 <>""'
Filters["no_false_negative"] = 'and n.等级<>"None" and n.解决办法 <>"" and n.误报="" '
Filters["high_risk"] = 'and not n.等级 in ["None","Low"] and n.解决办法 <>""'


def filter(TaskID, Scanner, Filter, TemplateFile, OutputFileName):
	if not OutputFileName:
		OutputFileName = TaskID
	outFile = OutputFileName + ".xlsx"
	
	wb = load_workbook(TemplateFile)
	sheet = wb.active
	
	headers = []
	for c in sheet.columns:
		headers.append(c[0].value)
		c[0].value = ""  # delete for quickly # empty headers, can't remove line use openpyxl
	condition = Filters[Filter]
	
	if Scanner:
		condition += ' and n.Scanner="%s"' % Scanner
	
	for v in DBO().enum_vul(TaskID, condition):
		line = []
		org_info = DBO().list_organization_structure(HostIP=v["IP"])[0]
		v.update(org_info)
		for c in headers:
			line.append(v[c])
		sheet.append(line)
	
	# auto line number
	if "TaskID" in headers:  # template may done have a TaskID
		c = headers.index("TaskID")
		i = 1
		for r in range(2, sheet.max_row):
			value = sheet.cell(column=c + 1, row=r + 1).value
			sheet.cell(column=c + 1, row=r + 1, value=value + str(i).zfill(5))
			i += 1
	
	# remove template data
	wb.remove(wb.get_sheet_by_name("Headers"))
	wb.save(outFile)
	print "[ Save as %s]\n" % outFile


if __name__ == '__main__':
	parser = argparse.ArgumentParser(description='''
    Export vulnerabilities data to user defined template file.
    1, Use key in '-c options' output, to make a template file,such as template_default.xlsx
    2, Use some filter supported to query data
    3, Default write to a file named: TaskID.xlsx
    ''', formatter_class=RawTextHelpFormatter, version="1.0")
	parser.add_argument("-c", dest="Columns", action='store_true',
	                    help="Print support data key, Use in excel Columns for user defined template.")
	parser.add_argument("-t", dest="TaskID", type=str,
	                    help="Unique TaskID for identify task and export.")
	parser.add_argument("-s", dest="Scanner", type=str, choices=["nessus", "nsfocus"],
	                    help="Select one scanner results.")
	parser.add_argument("-f", dest="Filter", type=str, choices=Filters.keys(), default="default",
	                    help="Filter some Data when query, Default filter vulnerabilities with risk. All will return all vulnerabilities")
	parser.add_argument("-m", dest="TemplateFile", type=file, default="template_default.xlsx",
	                    help="Export data to user predefined template file")
	parser.add_argument("-o", dest="OutputFileName", type=str,
	                    help="Store to Excel filename.")
	
	args = parser.parse_args()
	if args.TaskID and args.Filter and args.TemplateFile:
		filter(args.TaskID, args.Scanner, args.Filter, args.TemplateFile, args.OutputFileName)
	elif args.Columns:
		cols = MetaData().data.keys()
		[cols.append(x) for x in ["Project", "Department", "Application"]]
		cols.sort()
		for i in cols:
			print i
	else:
		parser.print_usage()
